from functools import lru_cache
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from app.config import get_settings

WEATHER_FIELD_NAMES = {
    "province": "省份",
    "city": "城市",
    "weather": "天气",
    "temperature": "温度",
    "winddirection": "风向",
    "windpower": "风力",
    "humidity": "湿度",
    "reporttime": "发布时间",
    "temperature_float": "温度数值",
    "humidity_float": "湿度数值",
    "date": "日期",
    "week": "星期",
    "dayweather": "白天天气",
    "nightweather": "夜间天气",
    "daytemp": "白天温度",
    "nighttemp": "夜间温度",
    "daywind": "白天风向",
    "nightwind": "夜间风向",
    "daypower": "白天风力",
    "nightpower": "夜间风力",
    "daytemp_float": "白天温度数值",
    "nighttemp_float": "夜间温度数值",
}
CITY_SUFFIXES = ("特别行政区", "自治州", "自治县", "地区", "盟", "市", "县", "区")
FORECAST_KEYWORDS = ("未来", "预报", "明天", "后天", "接下来", "近几天", "这几天", "一周", "all")


def get_weather(message: str | None = None) -> dict[str, Any]:
    settings = get_settings()
    if not settings.amap_weather_api_key:
        return _error_result(message, "未配置 AMAP_WEATHER_API_KEY，无法查询高德天气。")

    try:
        city = _match_city(message or "", settings.amap_adcode_path)
    except Exception as exc:
        return _error_result(message, f"读取城市编码表失败：{exc}")

    if not city:
        return _error_result(message, "未能从用户输入中识别城市或地区名称。")

    extensions = _resolve_extensions(message or "")
    params = {
        "city": city["adcode"],
        "key": settings.amap_weather_api_key,
        "extensions": extensions,
    }

    try:
        payload = _fetch_weather_payload(settings.amap_weather_base_url, params)
    except Exception as exc:
        return _error_result(message, f"高德天气请求失败：{exc}", city=city, extensions=extensions)

    if payload.get("status") != "1":
        error = payload.get("info") or payload.get("infocode") or "未知错误"
        return _error_result(message, f"高德天气接口返回异常：{error}", city=city, extensions=extensions)

    rows = _normalize_weather_payload(payload, extensions)
    formatted = _format_weather_rows(city["name"], extensions, rows)
    return {
        "tool": "weather",
        "implemented": True,
        "query": message,
        "city": city["name"],
        "extensions": extensions,
        "data": rows,
        "formatted": formatted,
    }


def _resolve_extensions(message: str) -> str:
    return "all" if any(keyword in message for keyword in FORECAST_KEYWORDS) else "base"


def _fetch_weather_payload(url: str, params: dict[str, str]) -> dict[str, Any]:
    try:
        with httpx.Client(timeout=15) as client:
            response = client.get(url, params=params)
            response.raise_for_status()
            return response.json()
    except httpx.TransportError as exc:
        if "SSL" not in str(exc).upper():
            raise

    with httpx.Client(timeout=15, verify=False) as client:
        response = client.get(url, params=params)
        response.raise_for_status()
        return response.json()


def _normalize_weather_payload(payload: dict[str, Any], extensions: str) -> list[dict[str, Any]]:
    if extensions == "all":
        rows: list[dict[str, Any]] = []
        for forecast in payload.get("forecasts", []):
            shared = _translate_fields(forecast, excluded={"adcode", "casts"})
            for cast in forecast.get("casts", []):
                rows.append({**shared, **_translate_fields(cast)})
        return rows
    return [_translate_fields(item, excluded={"adcode"}) for item in payload.get("lives", [])]


def _translate_fields(item: dict[str, Any], excluded: set[str] | None = None) -> dict[str, Any]:
    excluded = excluded or set()
    translated: dict[str, Any] = {}
    for key, value in item.items():
        if key in excluded:
            continue
        translated[WEATHER_FIELD_NAMES.get(key, key)] = value
    return translated


def _format_weather_rows(city_name: str, extensions: str, rows: list[dict[str, Any]]) -> str:
    weather_type = "天气预报" if extensions == "all" else "实时天气"
    if not rows:
        return f"高德{weather_type}结果：未获取到 {city_name} 的天气信息。"

    lines = [f"高德{weather_type}结果：{city_name}，共 {len(rows)} 条。"]
    for index, row in enumerate(rows, start=1):
        details = "，".join(f"{key}：{value}" for key, value in row.items())
        lines.append(f"{index}. {details}")
    return "\n".join(lines)


def _error_result(
    message: str | None,
    error: str,
    city: dict[str, str] | None = None,
    extensions: str | None = None,
) -> dict[str, Any]:
    return {
        "tool": "weather",
        "implemented": False,
        "query": message,
        "city": city.get("name") if city else None,
        "extensions": extensions,
        "data": [],
        "formatted": f"高德天气工具：{error}",
        "error": error,
    }


def _match_city(message: str, adcode_path: str) -> dict[str, str] | None:
    normalized_message = _normalize_name(message)
    if not normalized_message:
        return None

    candidates: list[tuple[int, int, dict[str, str]]] = []
    for city in _load_adcodes(adcode_path):
        names = {city["name"], _strip_city_suffix(city["name"])}
        names = {_normalize_name(name) for name in names if name}
        names = {name for name in names if len(name) >= 2}
        matched_names = [name for name in names if name and name in normalized_message]
        if matched_names:
            best_name = max(matched_names, key=len)
            candidates.append((_adcode_specificity(city["adcode"]), len(best_name), city))

    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return candidates[0][2]


def _adcode_specificity(adcode: str) -> int:
    if not adcode.endswith("00"):
        return 3
    if not adcode.endswith("0000"):
        return 2
    return 1


@lru_cache(maxsize=4)
def _load_adcodes(adcode_path: str) -> tuple[dict[str, str], ...]:
    path = Path(adcode_path)
    if not path.exists():
        raise FileNotFoundError(f"城市编码表不存在：{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        adcode = str(row[1] or "").strip()
        if not name or not adcode or name == "中国":
            continue
        rows.append({"name": name, "adcode": adcode})
    workbook.close()
    return tuple(rows)


def _strip_city_suffix(name: str) -> str:
    for suffix in CITY_SUFFIXES:
        if name.endswith(suffix) and len(name) > len(suffix):
            return name[: -len(suffix)]
    return name


def _normalize_name(value: str) -> str:
    return "".join(str(value).split())
