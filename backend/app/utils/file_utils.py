from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


ALLOWED_SUFFIXES = {".txt", ".md", ".pdf", ".docx", ".csv", ".xlsx", ".xls"}


def validate_upload(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        raise ValueError(f"不支持的文件类型: {suffix}")
    return suffix.replace(".", "") or "unknown"


def extract_text(path: str | Path) -> str:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return file_path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".docx":
        document = Document(str(file_path))
        return "\n".join(p.text for p in document.paragraphs if p.text.strip())
    if suffix == ".pdf":
        reader = PdfReader(str(file_path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix in {".xlsx", ".xls"}:
        workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        rows: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append(" | ".join("" if value is None else str(value) for value in row))
        return "\n".join(rows)
    return ""

